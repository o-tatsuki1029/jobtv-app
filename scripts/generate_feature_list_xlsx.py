"""
JOBTV 機能一覧 Excel 自動生成スクリプト

Usage:
    pip install openpyxl pyyaml
    python scripts/generate_feature_list_xlsx.py

Source:
    docs/feature-registry.yaml (機能台帳の正)

Output:
    docs/presentations/jobtv_feature_list.xlsx
"""

from collections import Counter
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parent.parent
REGISTRY_FILE = ROOT_DIR / "docs" / "feature-registry.yaml"
OUTPUT_DIR = ROOT_DIR / "docs" / "presentations"
OUTPUT_FILE = OUTPUT_DIR / "jobtv_feature_list.xlsx"

# ---------------------------------------------------------------------------
# Design tokens
# ---------------------------------------------------------------------------
DARK_NAVY_HEX = "1B2A4A"
ACCENT_BLUE_HEX = "2E86DE"
WHITE_HEX = "FFFFFF"
LIGHT_GRAY_HEX = "F0F2F5"

FONT_JP = "メイリオ"

HEADER_FONT = Font(name=FONT_JP, size=11, bold=True, color=WHITE_HEX)
HEADER_FILL = PatternFill(start_color=DARK_NAVY_HEX, end_color=DARK_NAVY_HEX, fill_type="solid")
NORMAL_FONT = Font(name=FONT_JP, size=10, color="2D3A4A")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
EVEN_ROW_FILL = PatternFill(start_color=LIGHT_GRAY_HEX, end_color=LIGHT_GRAY_HEX, fill_type="solid")

COLUMNS = ["No.", "アプリ", "大カテゴリ", "中カテゴリ", "機能名", "機能詳細", "対象ロール", "関連テーブル/API"]
COL_WIDTHS = [6, 16, 18, 20, 28, 60, 16, 40]

APP_ORDER = ["JOBTV", "Event System", "Agent Manager", "共通基盤"]


# ---------------------------------------------------------------------------
# Load YAML
# ---------------------------------------------------------------------------

def load_features():
    with open(REGISTRY_FILE, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data["features"]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _apply_header_style(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_row_style(ws, row_num, num_cols, is_even=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP_ALIGNMENT
        cell.border = THIN_BORDER
        if is_even:
            cell.fill = EVEN_ROW_FILL


def _set_col_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_main_sheet(wb, features):
    ws = wb.active
    ws.title = "機能一覧（全体）"
    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _apply_header_style(ws, 1, len(COLUMNS))

    for i, feat in enumerate(features):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=i + 1)
        ws.cell(row=row_num, column=2, value=feat["app"])
        ws.cell(row=row_num, column=3, value=feat["category"])
        ws.cell(row=row_num, column=4, value=feat["subcategory"])
        ws.cell(row=row_num, column=5, value=feat["name"])
        ws.cell(row=row_num, column=6, value=feat["detail"])
        ws.cell(row=row_num, column=7, value=feat["role"])
        ws.cell(row=row_num, column=8, value=feat["tables"])
        _apply_row_style(ws, row_num, len(COLUMNS), is_even=(i % 2 == 0))
        ws.cell(row=row_num, column=1).alignment = CENTER_ALIGNMENT

    _set_col_widths(ws, COL_WIDTHS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(features) + 1}"


def build_summary_sheet(wb, features):
    ws = wb.create_sheet("サマリ（アプリ×カテゴリ別件数）")

    app_cat_count = Counter()
    app_count = Counter()
    for feat in features:
        app_cat_count[(feat["app"], feat["category"])] += 1
        app_count[feat["app"]] += 1

    ws.cell(row=1, column=1, value="アプリ")
    ws.cell(row=1, column=2, value="大カテゴリ")
    ws.cell(row=1, column=3, value="機能数")
    _apply_header_style(ws, 1, 3)

    row = 2
    for app in APP_ORDER:
        categories = []
        seen = set()
        for f in features:
            if f["app"] == app and f["category"] not in seen:
                categories.append(f["category"])
                seen.add(f["category"])

        for cat in categories:
            ws.cell(row=row, column=1, value=app)
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=app_cat_count[(app, cat)])
            _apply_row_style(ws, row, 3, is_even=(row % 2 == 0))
            row += 1

        # Subtotal
        ws.cell(row=row, column=1, value=f"{app} 小計")
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value=app_count[app])
        subtotal_fill = PatternFill(start_color=ACCENT_BLUE_HEX, end_color=ACCENT_BLUE_HEX, fill_type="solid")
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name=FONT_JP, size=10, bold=True, color=WHITE_HEX)
            cell.fill = subtotal_fill
            cell.border = THIN_BORDER
        row += 1

    # Grand total
    ws.cell(row=row, column=1, value="総合計")
    ws.cell(row=row, column=2, value="")
    ws.cell(row=row, column=3, value=len(features))
    total_fill = PatternFill(start_color=DARK_NAVY_HEX, end_color=DARK_NAVY_HEX, fill_type="solid")
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_JP, size=11, bold=True, color=WHITE_HEX)
        cell.fill = total_fill
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 10
    ws.freeze_panes = "A2"


def build_app_sheets(wb, features):
    for app_name in APP_ORDER:
        app_features = [f for f in features if f["app"] == app_name]
        if not app_features:
            continue

        ws = wb.create_sheet(app_name[:31])
        ws.freeze_panes = "A2"

        headers = ["No.", "大カテゴリ", "中カテゴリ", "機能名", "機能詳細", "対象ロール", "関連テーブル/API"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(ws, 1, len(headers))

        for i, feat in enumerate(app_features):
            row_num = i + 2
            ws.cell(row=row_num, column=1, value=i + 1)
            ws.cell(row=row_num, column=2, value=feat["category"])
            ws.cell(row=row_num, column=3, value=feat["subcategory"])
            ws.cell(row=row_num, column=4, value=feat["name"])
            ws.cell(row=row_num, column=5, value=feat["detail"])
            ws.cell(row=row_num, column=6, value=feat["role"])
            ws.cell(row=row_num, column=7, value=feat["tables"])
            _apply_row_style(ws, row_num, len(headers), is_even=(i % 2 == 0))
            ws.cell(row=row_num, column=1).alignment = CENTER_ALIGNMENT

        _set_col_widths(ws, [6, 18, 20, 28, 60, 16, 40])
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(app_features) + 1}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    features = load_features()

    wb = Workbook()
    build_main_sheet(wb, features)
    build_summary_sheet(wb, features)
    build_app_sheets(wb, features)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_FILE))
    print(f"Source:    {REGISTRY_FILE}")
    print(f"Generated: {OUTPUT_FILE}")
    print(f"Total features: {len(features)}")


if __name__ == "__main__":
    main()
